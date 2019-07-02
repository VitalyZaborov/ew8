import openpyxl

INPUT = './Balance.xlsm'
OUTPUT = '../Assets/Script/GameParams.cs'
BASIC_TYPES = ('bool', 'int', 'uint', 'float', 'double', 'string', )

structFields = {}

def getData(cell, field, sheet, type):
	value = cell.value
	if value is None:
		return None
	if field in ('skills', 'chants'):
		array = value.split(',')
		res = 'new {'
		for i in range(0, len(array), 2):
			if i != 0:
				res += ', '
			skill = array[i]
			level = array[i + 1]
			res += skill + ' = ' + level
		res += '}'
		return res
	if type == 'DamageModifier':
		return 'new DamageModifier(' + str(value) + ')'
	if type not in BASIC_TYPES:
		return type + '.' + value
	if isinstance(value, str):
		return '"' + value + '"'
	return str(value) + ('f' if type == 'float' else '')


def getRow(row):
	fields = []
	for cell in row:
		if cell.value is None:
			break
		fields.append(str(cell.value))
	return fields

def getStruct(structName, superclassName, fields, types):
	result = '\n\n	public class ' + structName + (': ' + superclassName if superclassName is not None else '') + '{'

	superclassFields = structFields[superclassName] if superclassName is not None else []
	fields = [f for f in fields if f not in superclassFields]

	for field, type in zip(fields, types):
		result += '\n		public ' + type + ' ' + field + ';'
	result += '\n\n		public ' + structName + ' clone(){'
	result += '\n			return new ' + structName + '(){'
	rows = ['\n				' + field + ' = ' + field for field in fields]
	result += ','.join(rows)
	result += '\n			};\n		}'
	result += '\n	}'
	return result

def getParam(row, structName, fields, types):
	values = []
	key = None
	for i, field in enumerate(fields):
		value = getData(worksheet.cell(row=row, column=i + 1), field, structName, types[i])
		if i == 0:
			key = value
		values.append('\n			' + field + ' = ' + str(value));

	return '\n		{' + str(key) + ', new ' + structName + '{' + ','.join(values) + '\n		}}'


output = open(OUTPUT, 'w')
workbook = openpyxl.load_workbook(filename=INPUT, data_only=True)

output.write('using System.Collections;\nusing System.Collections.Generic;\n\npublic class GameParams{')

for worksheet in workbook.worksheets:
	if worksheet.title.startswith('_'):
		continue

	fields = getRow(worksheet.rows[0])
	types = getRow(worksheet.rows[1])
	structArray = worksheet.title.split(',')
	structName = structArray[0]
	superclassName = structArray[1] if len(structArray) > 1 else None
	varName = structName[0].lower() + structName[1:]
	structFields[structName] = fields

	output.write(getStruct(structName, superclassName, fields, types))

	typeDef = 'Dictionary<' + types[0] + ', ' + varName + '>'
	output.write('\n\n	public static ' + typeDef + ' ' + worksheet.title + ' = new ' + typeDef + '{')

	params = []
	i = 3
	while worksheet.cell(row=i, column=2).value is not None:
		params.append(getParam(i, structName, fields, types))
		i += 1

	output.write(','.join(params) + '\n	};')

output.write('\n}')
output.close()
