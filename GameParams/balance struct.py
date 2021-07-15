import openpyxl

INPUT = './Balance.xlsm'
OUTPUT = '../Assets/Script/GameParams.cs'
BASIC_TYPES = ('bool', 'int', 'uint', 'float', 'double', 'string', )

def getData(cell, field, sheet, fieldType):
	value = cell.value
	if value is None:
		return 'null'
	if field in ('skills', 'chants'):
		array = value.split(',')
		res = 'new {'
		for i in range(0, len(array), 2):
			if i != 0:
				res += ', '
			skill = array[i]
			level = array[i + 1]
			res += skill + ' = ' + level
		res += '}'
		return res
	if fieldType == 'DamageModifier':
		return 'new DamageModifier(' + str(value) + ')'
	if fieldType not in BASIC_TYPES:
		return fieldType + '.' + value
	if fieldType == 'string':
		return '"' + value + '"'
	return str(value) + ('f' if fieldType == 'float' else '')


def getRow(row):
	fields = []
	for cell in row:
		if cell.value is None:
			break
		fields.append(str(cell.value))
	return fields

def getStruct(structName, fields, types):
	result = '\n\n	public class ' + structName + '{'
	for field, type in zip(fields, types):
		result += '\n		public ' + type + ' ' + field + ';'
	result += '\n\n		public ' + structName + ' clone(){'
	result += '\n			return new ' + structName + '(){'
	rows = ['\n				' + field + ' = ' + field for field in fields]
	result += ','.join(rows)
	result += '\n			};\n		}'
	result += '\n	}'
	return result

def getParam(row, structName, fields, types):
	values = []
	key = None
	for i, field in enumerate(fields):
		value = getData(worksheet.cell(row=row, column=i + 1), field, structName, types[i])
		if i == 0:
			key = value
		values.append('\n			' + field + ' = ' + str(value))

	return '\n		{' + str(key) + ', new ' + structName + '{' + ','.join(values) + '\n		}}'


output = open(OUTPUT, 'w')
workbook = openpyxl.load_workbook(filename=INPUT, data_only=True)

output.write('using System.Collections;\nusing System.Collections.Generic;\n\npublic class GameParams{')

for worksheet in workbook.worksheets:
	if worksheet.title.startswith('_'):
		continue

	rows = [r for r in worksheet.rows]
	fields = getRow(rows[0])
	types = getRow(rows[1])
	structName = worksheet.title[0].upper() + worksheet.title[1:]

	output.write(getStruct(structName, fields, types))

	typeDef = 'Dictionary<' + types[0] + ', ' + structName + '>'
	output.write('\n\n	public static ' + typeDef + ' ' + worksheet.title + ' = new ' + typeDef + '{')

	params = []
	i = 3
	while worksheet.cell(row=i, column=2).value is not None:
		params.append(getParam(i, structName, fields, types))
		i += 1

	output.write(','.join(params) + '\n	};')

output.write('\n}')
output.close()
